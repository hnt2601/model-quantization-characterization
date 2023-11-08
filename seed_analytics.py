import torchvision as tv
import os
import openpyxl

from collections import defaultdict
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment
from pathlib import Path
from tqdm import tqdm
from common import MEAN, STD
from model_evaluate import evaluate


def extract_analytic_to_excel(workbook, data_list, title):
    worksheet = workbook.create_sheet(title=f"{title}%_quantized_layers")
    alignment = Alignment(horizontal="center", vertical="center")

    # Add headers to the sheet
    worksheet["A1"] = "Configuration Index"
    worksheet.merge_cells("A1:A2")

    worksheet["B1"] = ""
    worksheet["C1"] = ""
    worksheet.merge_cells("B1:C1")
    worksheet["B1"] = "Accuracy (%)"
    worksheet["B2"] = "FP32 Accuracy (%)"
    worksheet["C2"] = "INT8 Accuracy (%)"

    row_index = 3

    for i, data in data_list.items():
        for d in data:
            worksheet.cell(
                row=row_index, column=column_index_from_string("A"), value=str(i)
            )
            worksheet.cell(
                row=row_index,
                column=column_index_from_string("B"),
                value=d["accuracy"]["FP32"],
            )
            worksheet.cell(
                row=row_index,
                column=column_index_from_string("C"),
                value=d["accuracy"]["INT8"],
            )

            row_index += 1

    for row in worksheet.iter_rows(min_row=1, max_row=row_index, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = alignment


def main():
    report_path = "reports/mobilenetv2_seed_analysis.xlsx"
    if not os.path.exists(report_path):
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(report_path)

    nb_repetitions = 10
    nb_models = 10
    ratio_quantized = 94
    baseline_acc = 0.8493

    data_path = "/media/data/hoangnt/Projects/Datasets"
    directory_path = Path(f"pretrained/mobilenetv2/{ratio_quantized}")
    target_extension = ".onnx"
    matching_files = [
        file
        for file in directory_path.iterdir()
        if file.is_file() and file.suffix == target_extension
    ]

    batch_size = 1024

    transform = tv.transforms.Compose(
        [tv.transforms.ToTensor(), tv.transforms.Normalize(MEAN, STD)]
    )

    dataset = tv.datasets.CIFAR10(
        root=data_path,
        train=False,
        download=False,
        transform=transform,
    )

    pbar = tqdm(total=nb_models)

    data_list = defaultdict(list)

    for i, int8_model_path in enumerate(matching_files[:nb_models]):
        c = 0
        while c < nb_repetitions:
            int8_acc = evaluate(dataset, int8_model_path, batch_size)

            data = {
                "accuracy": {"FP32": baseline_acc, "INT8": round(int8_acc, 4)},
            }

            data_list[i].append(data)

            c += 1

        pbar.update(1)

    pbar.close()

    extract_analytic_to_excel(workbook, data_list, ratio_quantized)

    workbook.save(report_path)


if __name__ == "__main__":
    main()
